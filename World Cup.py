import requests
from bs4 import BeautifulSoup
import openpyxl
header = {"User-Agent":"Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 Edg/115.0.1901.200"
           }
response = requests.get("https://www.tzuqiu.cc/competitions/21/teamStats.do", headers = header)
html = response.text
soup = BeautifulSoup(html, "html.parser")
summary_table = soup.find("div", {"id": "summary1"}).find("table", class_="table")
attack_table = soup.find("div", {"id": "offensive1"}).find("table", class_="table")
defence_table = soup.find("div", {"id": "defensive1"}).find("table", class_="table")
pass_table = soup.find("div", {"id": "pass1"}).find("table", class_="table")
summary_rows = summary_table.find_all("tr")[1:]
attack_rows = attack_table.find_all("tr")[1:]
defence_rows = defence_table.find_all("tr")[1:]
pass_rows = pass_table.find_all("tr")[1:]
workbook = openpyxl.Workbook()
summary_sheet = workbook.active
summary_sheet.title = "Summary Data"
attack_sheet = workbook.create_sheet(title="Attack Data")
defence_sheet = workbook.create_sheet(title="Defence Data")
pass_sheet = workbook.create_sheet(title="Pass Data")
custom_summary_header = ["Team Name in Chinese", "Shots", "Yellow Cards", "Red Cards", "Possession %", "Pass %", "Big-Chances", "Aerial Success", "Overall Rating"]
custom_attack_header = ["Team Name in Chinese", "Goals", "Shots", "Shots on Target", "Big-Chance", "Big-Chance Conversion", "Dribbles", "Fouled", "Offsides", "Rating"]
custom_defence_header = ["Team Name in Chinese", "Goals Conceded", "Shots Conceded", "Tackles", "Interception", "Clearances", "Offsides", "Fouls", "Errors", "Rating"]
custom_pass_header = ["Team Name in Chinese", "Assists", "Key Passes", "Possession", "Passes", "Pass %", "Final Third Pass %", "Rating"]
summary_sheet.append(custom_summary_header)
attack_sheet.append(custom_attack_header)
defence_sheet.append(custom_defence_header)
pass_sheet.append(custom_pass_header)
for row in summary_rows:
    cells = row.find_all("td")
    team_name = cells[1].a.get_text(strip=True)
    shots = cells[2].get_text(strip=True)
    yellow_cards = cells[3].find(class_="yel-card").get_text(strip=True)
    red_cards = cells[3].find(class_="red-card").get_text(strip=True)
    possession = cells[4].get_text(strip=True)
    pass_success = cells[5].get_text(strip=True)
    big_chances = cells[6].get_text(strip=True)
    aerial_success = cells[7].get_text(strip=True)
    rating = cells[8].get_text(strip=True)
    row_data = [team_name, shots, yellow_cards, red_cards, possession, pass_success, big_chances,
                aerial_success, rating]
    summary_sheet.append(row_data)

for row in attack_rows:
    cells = row.find_all("td")
    team_name = cells[1].a.get_text()
    goals = cells[2].get_text()
    shots = cells[3].get_text()
    shots_on_target = cells[4].get_text()
    big_chances = cells[5].get_text()
    big_chances_conversion = cells[6].get_text()
    dribbles = cells[7].get_text()
    fouled = cells[8].get_text()
    offsides = cells[9].get_text()
    rating = cells[10].get_text()
    row_data = [team_name, goals, shots, shots_on_target, big_chances, big_chances_conversion,dribbles, fouled, offsides, rating]
    attack_sheet.append(row_data)

for row in defence_rows:
    cells = row.find_all("td")
    team_name = cells[1].a.get_text(strip=True)
    goals_conceded = cells[2].get_text(strip=True)
    shots_conceded = cells[3].get_text(strip=True)
    tackles = cells[4].get_text(strip=True)
    interception = cells[5].get_text(strip=True)
    clearances = cells[6].get_text(strip=True)
    offsides = cells[7].get_text(strip=True)
    fouls = cells[8].get_text(strip=True)
    errors = cells[9].get_text(strip=True)
    rating = cells[10].get_text(strip=True)
    row_data = [team_name, goals_conceded, shots_conceded, tackles, interception , clearances, offsides, fouls, errors, rating]
    defence_sheet.append(row_data)

for row in pass_rows:
    cells = row.find_all("td")
    team_name = cells[1].a.get_text(strip=True)
    assists = cells[2].get_text(strip=True)
    key_passes = cells[3].get_text(strip=True)
    possession = cells[4].get_text(strip=True)
    passes = cells[5].get_text(strip=True)
    pass_rate = cells[6].get_text(strip=True)
    final_third_pass = cells[7].get_text(strip=True)
    rating = cells[8].get_text(strip=True)
    row_data = [team_name,assists , key_passes , possession , passes, pass_rate, final_third_pass, rating]
    pass_sheet.append(row_data)

workbook.save("World_Cup_Team_Data.xlsx")
